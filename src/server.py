# server.py

import socket               # For network communication between client and server
import threading            # To handle multiple clients concurrently
import os                   # For interacting with the operating system (e.g., file paths)
import glob                 # To find all the pathnames matching a specified pattern
import PyPDF2               # To read and extract text from PDF files
import openpyxl             # To read and manipulate Excel (.xlsx) files
import re                   # For regular expression operations
from bs4 import BeautifulSoup  # To parse and extract data from HTML files
import logging              # For logging server activities and errors

# Configure logging to output debug information to 'server.log'
logging.basicConfig(
    filename='server.log',              # Log file name
    level=logging.DEBUG,                # Logging level set to DEBUG for detailed output
    format='%(asctime)s - %(levelname)s - %(message)s'  # Log message format
)

# Define server constants
HOST = '127.0.0.1'      # Server's IP address (localhost)
PORT = 12345            # Port number where the server listens for connections
ENDING_MSG = "q"        # Special message indicating client wants to terminate the connection

# Mapping of file extensions to their corresponding directories
FOLDERS = {
    ".txt": "data/txt/",
    ".pdf": "data/pdf/",
    ".html": "data/html/",
    ".xlsx": "data/excel/"
}

def parse_keywords(keyword):
    """
    Parses the input keyword string to identify logical operators (AND/OR) or regex patterns.
    
    Args:
        keyword (str): The keyword string input by the user.
        
    Returns:
        tuple: A tuple containing a list of keywords, the operator used ('OR', 'AND', or 'SINGLE'),
               and a boolean indicating if the keyword is a regex pattern.
    """
    # Check if the keyword contains the 'OR' operator
    if " OR " in keyword:
        # Split the keyword string by ' OR ', strip whitespace, and return with operator 'OR'
        return [kw.strip() for kw in keyword.split(" OR ")], "OR", False
    # Check if the keyword contains the 'AND' operator
    elif " AND " in keyword:
        # Split the keyword string by ' AND ', strip whitespace, and return with operator 'AND'
        return [kw.strip() for kw in keyword.split(" AND ")], "AND", False

    # If no logical operator is found, check if the keyword is a valid regex pattern
    try:
        re.compile(keyword)  # Attempt to compile the keyword as a regex
        is_regex = True      # If successful, it's a regex pattern
    except re.error:
        is_regex = False     # If compilation fails, it's not a regex

    # Return the keyword as a single-item list, operator 'SINGLE', and regex flag
    return [keyword.strip()], "SINGLE", is_regex

def matches_with_operator(line, keywords, operator, is_regex):
    """
    Determines if a given line matches the search criteria based on keywords and the operator.
    
    Args:
        line (str): The line of text to be evaluated.
        keywords (list): List of keywords to search for.
        operator (str): The logical operator ('OR', 'AND', or 'SINGLE').
        is_regex (bool): Indicates whether the keywords should be treated as regex patterns.
        
    Returns:
        bool: True if the line matches the search criteria, False otherwise.
    """
    if operator == "OR":
        # For 'OR' operator, return True if any keyword matches the line
        return any(re.search(kw, line) if is_regex else kw in line for kw in keywords)
    elif operator == "AND":
        # For 'AND' operator, return True only if all keywords match the line
        return all(re.search(kw, line) if is_regex else kw in line for kw in keywords)
    elif operator == "SINGLE":
        # For 'SINGLE' operator, return True if the single keyword matches the line
        return re.search(keywords[0], line) if is_regex else keywords[0] in line
    return False  # Default to False if operator is unrecognized

def search_txt(file_path, keyword):
    """
    Searches for keywords within a TXT file.
    
    Args:
        file_path (str): The path to the TXT file.
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found.
    """
    results = []              # Initialize an empty list to store search results
    count = 1                 # Initialize a counter for numbering results
    keywords, operator, is_regex = parse_keywords(keyword)  # Parse the keyword string

    logging.debug(f"Searching in file: {file_path} with keywords={keywords}, operator={operator}, is_regex={is_regex}")
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()  # Read all lines from the TXT file
            for line_num, line in enumerate(lines, start=1):
                if matches_with_operator(line, keywords, operator, is_regex):
                    # If the line matches the search criteria, log the match and add to results
                    logging.debug(f"Match found in {file_path} on line {line_num}: {line.strip()}")
                    results.append(f"{count} - {os.path.basename(file_path)} - Line {line_num}: {line.strip()}")
                    count += 1
    except Exception as e:
        # Log any errors encountered while reading the TXT file
        logging.error(f"Error reading TXT file {file_path}: {e}")

    if not results:
        # Log if no matches were found in the file
        logging.debug(f"No matches found in file: {file_path}")
    return results  # Return the list of results

def search_pdf(file_path, keyword):
    """
    Searches for keywords within a PDF file.
    
    Args:
        file_path (str): The path to the PDF file.
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found.
    """
    results = []              # Initialize an empty list to store search results
    count = 1                 # Initialize a counter for numbering results
    keywords, operator, is_regex = parse_keywords(keyword)  # Parse the keyword string

    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)  # Create a PDF reader object
            for page_num, page in enumerate(reader.pages, start=1):
                text = page.extract_text()    # Extract text from the current page
                if not text:
                    continue  # Skip if no text is found on the page
                lines = text.split('\n')       # Split the extracted text into lines
                for line_num, line in enumerate(lines, start=1):
                    if matches_with_operator(line, keywords, operator, is_regex):
                        # If the line matches the search criteria, add to results
                        results.append(f"{count} - {os.path.basename(file_path)} - Page {page_num}, Line {line_num}: {line.strip()}")
                        count += 1
    except Exception as e:
        # Log any errors encountered while reading the PDF file
        logging.error(f"Error reading PDF file {file_path}: {e}")

    return results  # Return the list of results

def search_html(file_path, keyword):
    """
    Searches for keywords within an HTML file.
    
    Args:
        file_path (str): The path to the HTML file.
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found.
    """
    results = []              # Initialize an empty list to store search results
    count = 1                 # Initialize a counter for numbering results
    keywords, operator, is_regex = parse_keywords(keyword)  # Parse the keyword string

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file, 'html.parser')  # Parse the HTML content
            for tag in soup.find_all(text=True):       # Iterate through all text elements
                tag_text = tag.strip()                 # Remove leading/trailing whitespace
                if not tag_text:
                    continue  # Skip empty text
                if matches_with_operator(tag_text, keywords, operator, is_regex):
                    # If the text matches the search criteria, add to results
                    results.append(f"{count} - {os.path.basename(file_path)} - {tag_text}")
                    count += 1
    except Exception as e:
        # Log any errors encountered while reading the HTML file
        logging.error(f"Error reading HTML file {file_path}: {e}")

    return results  # Return the list of results

def search_xlsx(file_path, keyword):
    """
    Searches for keywords within an Excel (XLSX) file.
    
    Args:
        file_path (str): The path to the XLSX file.
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found.
    """
    results = []              # Initialize an empty list to store search results
    count = 1                 # Initialize a counter for numbering results
    keywords, operator, is_regex = parse_keywords(keyword)  # Parse the keyword string

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)  # Load the Excel workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  # Select the current sheet
            for row in sheet.iter_rows():
                for cell in row:
                    # Convert cell value to string if it's not None, else use empty string
                    cell_value = str(cell.value) if cell.value else ""
                    if matches_with_operator(cell_value, keywords, operator, is_regex):
                        # If the cell matches the search criteria, add to results
                        results.append(f"{count} - {os.path.basename(file_path)} - Sheet {sheet_name}, Cell ({cell.row}, {cell.column}): {cell_value}")
                        count += 1
    except Exception as e:
        # Log any errors encountered while reading the Excel file
        logging.error(f"Error reading Excel file {file_path}: {e}")

    return results  # Return the list of results

def search_all_files(folder, extension, keyword):
    """
    Searches for keywords within all files of a specific type in a given folder.
    
    Args:
        folder (str): The directory containing the files.
        extension (str): The file extension to search for.
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found or a message if no matches.
    """
    results = []  # Initialize an empty list to store search results
    for file_path in glob.glob(os.path.join(folder, f"*{extension}")):
        # Iterate through all files in the folder matching the extension
        if extension == ".txt":
            results.extend(search_txt(file_path, keyword))     # Search in TXT files
        elif extension == ".pdf":
            results.extend(search_pdf(file_path, keyword))     # Search in PDF files
        elif extension == ".html":
            results.extend(search_html(file_path, keyword))    # Search in HTML files
        elif extension == ".xlsx":
            results.extend(search_xlsx(file_path, keyword))    # Search in Excel files
    return results if results else ["No matches found in any file."]  # Return results or a default message

def search_all_file_types(keyword):
    """
    Searches for keywords across all defined file types.
    
    Args:
        keyword (str): The keyword or regex pattern to search for.
        
    Returns:
        list: A list of formatted strings indicating where matches were found or a message if no matches.
    """
    results = []  # Initialize an empty list to store search results
    for extension, folder in FOLDERS.items():
        # Iterate through each file type and its corresponding folder
        results.extend(search_all_files(folder, extension, keyword))  # Perform search
    return results if results else ["No matches found in any file."]  # Return results or a default message

def handle_search(search_target, keyword, file_extension):
    """
    Handles the search logic based on the target (ALL or specific file) and file extensions provided.
    
    Args:
        search_target (str): "ALL" to search all files or the name of a specific file.
        keyword (str): The keyword or regex pattern to search for.
        file_extension (str): Comma-separated file extensions to include in the search.
        
    Returns:
        list: A list of formatted strings indicating where matches were found or error messages.
    """
    # Split multiple extensions separated by commas and remove any extra whitespace
    extensions = [ext.strip() for ext in file_extension.split(',') if ext.strip()] if file_extension else []

    if search_target.upper() == "ALL":
        # If the search target is "ALL", search across all or specified file types
        if not extensions:
            # If no specific extensions are provided, search across all defined file types
            return search_all_file_types(keyword)
        else:
            # If specific extensions are provided, search only within those types
            results = []
            for ext in extensions:
                folder = FOLDERS.get(ext)
                if not folder:
                    # If the file type is unsupported, add an error message to results
                    results.append(f"Unsupported file type: {ext}")
                else:
                    # Perform search within the specified folder and extension
                    res = search_all_files(folder, ext, keyword)
                    if not res or (len(res) == 1 and res[0].startswith("No matches")):
                        # If no results found for this extension, add a corresponding message
                        res = [f"No matches found for file type: {ext}."]
                    results.extend(res)
            return results  # Return aggregated results
    else:
        # If the search target is a specific file
        if len(extensions) > 1:
            # If multiple extensions are provided for a single file search, return an error
            return ["Error: Multiple extensions provided for a single file search."]
        ext = extensions[0] if extensions else ""  # Get the single extension if available
        folder = FOLDERS.get(ext)
        if not folder:
            # If the file type is unsupported, return an error message
            return [f"Unsupported file type: {ext}"]
        file_path = os.path.join(folder, search_target)  # Construct the full file path
        if os.path.exists(file_path):
            # If the file exists, perform the appropriate search based on its extension
            if ext == ".txt":
                return search_txt(file_path, keyword) or ["No matches found."]
            elif ext == ".pdf":
                return search_pdf(file_path, keyword) or ["No matches found."]
            elif ext == ".html":
                return search_html(file_path, keyword) or ["No matches found."]
            elif ext == ".xlsx":
                return search_xlsx(file_path, keyword) or ["No matches found."]
        else:
            # If the file does not exist, return an error message
            return [f"File not found: {file_path}"]

def handle_client(client_socket):
    """
    Handles communication with a connected client.
    
    Args:
        client_socket (socket.socket): The client's socket connection.
    """
    try:
        while True:
            # Receive message from the client (max 1024 bytes)
            client_msg = client_socket.recv(1024).decode('utf-8')
            logging.debug(f"Received message: {client_msg}")
            if client_msg == ENDING_MSG:
                # If the client sends the termination message, break the loop to close connection
                logging.debug("Termination message received. Closing connection.")
                break
            try:
                # Parse the received message assuming format: <search_target>|<keyword>|<file_extension>
                search_target, keyword, file_extension = client_msg.split("|")
                # Strip any leading/trailing whitespace from each component
                search_target = search_target.strip()
                keyword = keyword.strip()
                file_extension = file_extension.strip()
                logging.debug(f"Parsed: search_target={search_target}, keyword={keyword}, file_extension={file_extension}")
                # Handle the search based on parsed components
                results = handle_search(search_target, keyword, file_extension)
                # Join the list of results into a single string separated by newlines
                response = "\n".join(results)
            except ValueError:
                # If message format is incorrect, prepare an error message
                response = "Invalid format. Use: <search_target>|<keyword>|<file_extension>"
            # Send the response back to the client encoded in UTF-8
            client_socket.send(response.encode('utf-8'))
    except Exception as e:
        # Log any unexpected errors while handling the client
        logging.error(f"Error handling client: {e}")
    finally:
        # Ensure the client socket is closed when done
        client_socket.close()
        logging.debug("Client connection closed.")

def main():
    """
    The main function to start the server, listen for incoming connections,
    and handle client requests.
    """
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as server:
        try:
            server.bind((HOST, PORT))      # Bind the server to the specified host and port
            server.listen(5)               # Listen for incoming connections with a backlog of 5
            print(f"[Server] Running on {HOST}:{PORT}. Press 'q' to stop.")
            logging.info(f"Server started on {HOST}:{PORT}")

            def stop_server():
                """
                Function to stop the server when the user inputs 'q'.
                Runs in a separate daemon thread.
                """
                while True:
                    if input().strip().lower() == 'q':
                        print("[Server] Shutting down.")
                        logging.info("Server shutdown initiated by user.")
                        os._exit(0)  # Forcefully exit the program

            # Start the stop_server function in a separate daemon thread
            threading.Thread(target=stop_server, daemon=True).start()

            while True:
                # Accept an incoming client connection
                client_socket, addr = server.accept()
                logging.info(f"Accepted connection from {addr}")
                # Start a new thread to handle the connected client
                threading.Thread(target=handle_client, args=(client_socket,)).start()
        except Exception as e:
            # Log any errors that occur during server setup or execution
            logging.error(f"Server error: {e}")
            print(f"[Server] Error: {e}")

if __name__ == "__main__":
    main()  # Execute the main function when the script is run directly
